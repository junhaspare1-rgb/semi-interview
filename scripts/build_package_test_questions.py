import json
import sys
from pathlib import Path

from openpyxl import load_workbook


DIFFICULTY_ALIASES = {
    "실무": "실전",
}
REQUIRED_HEADERS = {
    "No.": "id",
    "카테고리": "category",
    "질문": "question",
    "2분 모범 답안 Script (600~650자)": "answer",
    "40초 Script (280~320자)": "shortAnswer",
    "답변 Keyword": "keywords",
    "난이도": "difficulty",
}


def normalize_text(value):
    return " ".join(str(value or "").split())


def normalize_difficulty(value):
    text = normalize_text(value)
    return DIFFICULTY_ALIASES.get(text, text or "입문")


def split_keywords(value):
    text = str(value or "")
    chunks = []
    for line in text.splitlines():
        chunks.extend(line.split(","))
    return [chunk.strip() for chunk in chunks if chunk.strip()]


def header_map(sheet):
    headers = [normalize_text(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Missing required headers: {', '.join(missing)}")
    return {header: headers.index(header) + 1 for header in REQUIRED_HEADERS}


def build_records(source):
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    columns = header_map(sheet)

    records = []
    for row in range(2, sheet.max_row + 1):
        question = normalize_text(sheet.cell(row, columns["질문"]).value)
        if not question:
            continue

        id_value = sheet.cell(row, columns["No."]).value
        difficulty = normalize_difficulty(sheet.cell(row, columns["난이도"]).value)

        records.append(
            {
                "id": int(id_value),
                "jobRole": "Package & Test",
                "category": normalize_text(sheet.cell(row, columns["카테고리"]).value),
                "group": "other",
                "difficulty": difficulty,
                "question": question,
                "answer": normalize_text(sheet.cell(row, columns["2분 모범 답안 Script (600~650자)"]).value),
                "keywords": split_keywords(sheet.cell(row, columns["답변 Keyword"]).value),
                "active": True,
                "estimatedAnswerMinutes": 2,
                "shortAnswer": normalize_text(sheet.cell(row, columns["40초 Script (280~320자)"]).value),
            }
        )

    records.sort(key=lambda record: record["id"])
    ids = [record["id"] for record in records]
    if len(ids) != len(set(ids)):
        raise ValueError("Duplicate question No. values found.")

    return records


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python scripts/build_package_test_questions.py <source.xlsx> [output_dir]")

    source = Path(sys.argv[1])
    output_dir = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path("data")
    output_dir.mkdir(parents=True, exist_ok=True)

    records = build_records(source)
    json_text = json.dumps(records, ensure_ascii=False, indent=2)

    json_path = output_dir / "package-test-questions.json"
    js_path = output_dir / "package-test-questions.js"
    json_path.write_text(json_text + "\n", encoding="utf-8")
    js_path.write_text(
        "window.BANMYEONPPU_PACKAGE_TEST_QUESTIONS = "
        + json_text
        + ";\n",
        encoding="utf-8",
    )

    difficulty_counts = {}
    for record in records:
        difficulty_counts[record["difficulty"]] = difficulty_counts.get(record["difficulty"], 0) + 1

    active_count = sum(1 for record in records if record["difficulty"] != "지엽")
    print(f"Wrote {len(records)} records to {json_path} and {js_path}")
    print(f"Active question count excluding 지엽: {active_count}")
    print(difficulty_counts)


if __name__ == "__main__":
    main()
