import json
import sys
from pathlib import Path

from openpyxl import load_workbook


EXCLUDED_CATEGORIES = {"기타"}
REQUIRED_HEADERS = {
    "No.": "id",
    "카테고리": "category",
    "질문": "question",
    "권장 답변": "recommendedAnswer",
    "지양 답변": "avoidAnswer",
}


def normalize_text(value):
    return " ".join(str(value or "").split())


def header_map(sheet):
    headers = [normalize_text(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Missing required headers: {', '.join(missing)}")
    return {header: headers.index(header) + 1 for header in REQUIRED_HEADERS}


def build_records(source):
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["질문목록"] if "질문목록" in workbook.sheetnames else workbook.worksheets[0]
    columns = header_map(sheet)

    records = []
    for row in range(2, sheet.max_row + 1):
        question = normalize_text(sheet.cell(row, columns["질문"]).value)
        category = normalize_text(sheet.cell(row, columns["카테고리"]).value)
        if not question or category in EXCLUDED_CATEGORIES:
            continue

        id_value = normalize_text(sheet.cell(row, columns["No."]).value)
        records.append(
            {
                "id": id_value,
                "jobRole": "인성 면접",
                "category": category or "미분류",
                "group": "other",
                "difficulty": "입문",
                "question": question,
                "questionType": "personality",
                "recommendedAnswer": normalize_text(sheet.cell(row, columns["권장 답변"]).value),
                "avoidAnswer": normalize_text(sheet.cell(row, columns["지양 답변"]).value),
                "answer": normalize_text(sheet.cell(row, columns["권장 답변"]).value),
                "keywords": [],
                "active": True,
                "estimatedAnswerMinutes": None,
                "shortAnswer": "",
            }
        )

    ids = [record["id"] for record in records]
    if len(ids) != len(set(ids)):
        raise ValueError("Duplicate question No. values found.")

    return records


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python scripts/build_personality_questions.py <source.xlsx> [output_dir]")

    source = Path(sys.argv[1])
    output_dir = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path("data")
    output_dir.mkdir(parents=True, exist_ok=True)

    records = build_records(source)
    json_text = json.dumps(records, ensure_ascii=False, indent=2)

    json_path = output_dir / "personality-questions.json"
    js_path = output_dir / "personality-questions.js"
    json_path.write_text(json_text + "\n", encoding="utf-8")
    js_path.write_text(
        "window.BANMYEONPPU_PERSONALITY_QUESTIONS = "
        + json_text
        + ";\n",
        encoding="utf-8",
    )

    category_counts = {}
    for record in records:
        category_counts[record["category"]] = category_counts.get(record["category"], 0) + 1

    print(f"Wrote {len(records)} records to {json_path} and {js_path}")
    print(category_counts)


if __name__ == "__main__":
    main()
