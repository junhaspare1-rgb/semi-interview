import json
import sys
from pathlib import Path

from openpyxl import load_workbook


MAIN_CATEGORIES = {"포토(Lithography)", "식각(Etch)", "증착(Deposition)"}
DIFFICULTY_ALIASES = {
    "실무": "실전",
}


def normalize_difficulty(value):
    text = str(value or "").strip()
    return DIFFICULTY_ALIASES.get(text, text)


def split_keywords(value):
    return [part.strip() for part in str(value or "").split(",") if part.strip()]


def split_tail_questions(value):
    text = str(value or "").strip()
    if not text:
        return []
    normalized = text.replace("；", ";")
    parts = []
    for chunk in normalized.split("\n"):
        parts.extend(chunk.split(";"))
    return [part.strip() for part in parts if part.strip()]


def build_records(source):
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]

    def column(name):
        return headers.index(name) + 1

    no_col = column("No.")
    category_col = column("카테고리")
    question_col = column("면접 질문")
    answer_col = column("모범 답안 Script")
    keywords_col = column("핵심 키워드")
    difficulty_col = column("난이도")
    tail_questions_col = headers.index("꼬리질문") + 1 if "꼬리질문" in headers else None

    records = []
    for row in range(2, sheet.max_row + 1):
        question = sheet.cell(row, question_col).value
        if not question:
            continue

        category = str(sheet.cell(row, category_col).value or "").strip()
        difficulty = normalize_difficulty(sheet.cell(row, difficulty_col).value)
        records.append(
            {
                "id": int(sheet.cell(row, no_col).value),
                "jobRole": "공정기술",
                "category": category,
                "group": "main" if category in MAIN_CATEGORIES else "other",
                "difficulty": difficulty,
                "question": str(question).strip(),
                "answer": str(sheet.cell(row, answer_col).value or "").strip(),
                "keywords": split_keywords(sheet.cell(row, keywords_col).value),
                "tailQuestions": split_tail_questions(sheet.cell(row, tail_questions_col).value)
                if tail_questions_col
                else [],
                "active": difficulty != "지엽",
            }
        )

    return records


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python scripts/build_process_questions.py <source.xlsx> [output_dir]")

    source = Path(sys.argv[1])
    output_dir = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path("data")
    output_dir.mkdir(parents=True, exist_ok=True)

    records = build_records(source)
    json_text = json.dumps(records, ensure_ascii=False, indent=2)

    json_path = output_dir / "process-questions.json"
    js_path = output_dir / "process-questions.js"
    json_path.write_text(json_text + "\n", encoding="utf-8")
    js_path.write_text(
        "window.BANMYEONPPU_PROCESS_QUESTIONS = "
        + json_text
        + ";\n",
        encoding="utf-8",
    )

    counts = {}
    for record in records:
        if record["active"]:
            counts[record["difficulty"]] = counts.get(record["difficulty"], 0) + 1

    print(f"Wrote {len(records)} records to {json_path} and {js_path}")
    print(counts)


if __name__ == "__main__":
    main()
